#!/usr/bin/env python3
"""
timeline_style.py — keep the timeline workbook's cell formatting uniform.

Why: rows appended programmatically (e.g. the block-U "add events" pass) arrive
with default styling (Sans 10, no wrap, bottom-aligned), which breaks the
sheet's look and stops LibreOffice from auto-fitting row heights. Rather than
hard-coding a style, this tool treats the FIRST DATA ROW (row 2) of each sheet
as the canonical template, per column, and applies it to every row below.

So: restyle row 2 by hand (font, alignment, wrap, borders) and run
`python3 tools/timeline_sync.py --refresh-only` — the whole sheet follows.
timeline_sync.py calls normalize() automatically on every refresh, so newly
appended rows can never stay unstyled.

Deliberately untouched: cell VALUES, row heights (LibreOffice auto-fits wrapped
rows without an explicit height; manual heights are respected), column widths,
the header row, and the Reader sheet (rebuilt wholesale by build_reader.py).

Usage:
  python3 tools/timeline_style.py          # normalize + report
  python3 tools/timeline_style.py --check  # report only, change nothing
"""
import os, sys, argparse
from copy import copy

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "analysis", "IceCapade_timeline.xlsx")

SKIP_SHEETS = ("Reader",)   # matched as a prefix of the sheet title
TEMPLATE_ROW = 2            # first data row = the canonical style, per column


def _style_key(c):
    """Comparable fingerprint of the styling we manage."""
    a, f = c.alignment, c.font
    return (a.horizontal, a.vertical, a.wrap_text,
            f.name, f.size, f.bold, f.italic,
            c.number_format)


def normalize(check_only=False):
    """Apply row-2 styling down each sheet. Returns number of cells restyled."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)
    changed = 0
    for ws in wb.worksheets:
        if ws.title.startswith(SKIP_SHEETS) or ws.max_row <= TEMPLATE_ROW:
            continue
        for col in range(1, ws.max_column + 1):
            src = ws.cell(TEMPLATE_ROW, col)
            key = _style_key(src)
            for row in range(TEMPLATE_ROW + 1, ws.max_row + 1):
                dst = ws.cell(row, col)
                if _style_key(dst) == key:
                    continue
                changed += 1
                if check_only:
                    continue
                dst.font          = copy(src.font)
                dst.alignment     = copy(src.alignment)
                dst.border        = copy(src.border)
                dst.fill          = copy(src.fill)
                dst.protection    = copy(src.protection)
                dst.number_format = src.number_format
    if changed and not check_only:
        wb.save(XLSX)
    return changed


def main():
    ap = argparse.ArgumentParser(description="Normalize timeline workbook cell styles to each sheet's row-2 template.")
    ap.add_argument("--check", action="store_true", help="report drift only; do not write")
    args = ap.parse_args()
    n = normalize(check_only=args.check)
    if args.check:
        print(f"style drift: {n} cell(s) differ from the row-2 template."
              if n else "styles uniform: every data row matches its row-2 template.")
        sys.exit(1 if n else 0)
    print(f"restyled {n} cell(s) to the row-2 template." if n
          else "styles already uniform; nothing to do.")


if __name__ == "__main__":
    main()
