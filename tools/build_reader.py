#!/usr/bin/env python3
"""
build_reader.py — (re)generate the pre-rendered "Reader" sheet in the timeline workbook.

The Reader is a read-only, one-card-per-entry view of the two chronology sheets,
rendered as plain text (no formulas, no controls) so it displays identically in
Excel, LibreOffice, and Gnumeric. Jump to an entry with Ctrl+F (e.g. "S34").

Because it is static, it does NOT track edits to the data sheets — rerun this
(or `tools/timeline_sync.py`, which calls it) after any workbook change.

Usage:
  python3 tools/build_reader.py
"""
import os, math, warnings
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX  = os.path.join(ROOT, "analysis", "IceCapade_timeline.xlsx")

CHRON_PREFIXES = ("Historical", "Slow Mend")   # sheets to render as cards

# nicer section-banner labels (fallback: the sheet title, uppercased)
DIVIDER_LABELS = {
    "Historical (2008-2055)": "HISTORICAL  (2008–2055)",
    "Slow Mend (Yr 0-100)":   "SLOW MEND  (Year 0–100)",
}

# ---- 2x-scaled styling ----
NAVY, BAND, HINT = "1F3864", "2E5496", "808080"
F_TITLE  = Font(bold=True, size=30, color="FFFFFF")
F_BAND   = Font(bold=True, size=24, color="FFFFFF")
F_DIV    = Font(bold=True, size=26, color="FFFFFF")
F_LABEL  = Font(bold=True, size=20, color=NAVY)
F_VALUE  = Font(size=22, color="222222")
F_HINT   = Font(italic=True, size=20, color=HINT)
TOP_L    = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR_L    = Alignment(horizontal="left", vertical="center")
UNDER    = Border(bottom=Side(style="thin", color="C9C9C9"))
CPL, LH  = 98, 30          # chars/line, per-line height (both 2x baseline)


def _rows_of(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 11)]
        if v[0] and str(v[0]).strip():
            out.append(["" if x is None else str(x) for x in v])
    return out


def build_reader(xlsx_path=XLSX):
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(xlsx_path)
    chron = [ws for ws in wb.worksheets if ws.title.startswith(CHRON_PREFIXES)]

    if "Reader" in wb.sheetnames:
        del wb["Reader"]
    rd = wb.create_sheet("Reader", 0)
    rd.column_dimensions["A"].width = 30
    rd.column_dimensions["B"].width = 208

    state = {"r": 1}

    def band(text):
        r = state["r"]; rd.merge_cells(f"A{r}:B{r}")
        c = rd.cell(r, 1); c.value = text; c.font = F_BAND
        c.fill = PatternFill("solid", fgColor=BAND); c.alignment = CTR_L
        rd.row_dimensions[r].height = 40; state["r"] += 1

    def field(label, value, big=False, small=False):
        if value is None or str(value).strip() == "":
            return
        r = state["r"]
        a = rd.cell(r, 1); a.value = label; a.font = F_LABEL; a.alignment = TOP_L; a.border = UNDER
        b = rd.cell(r, 2); b.value = value; b.font = F_VALUE; b.alignment = TOP_L; b.border = UNDER
        lines = max(1, math.ceil(len(str(value)) / CPL))
        if big:     h = min(max(2 * LH, lines * LH + 12), 520)
        elif small: h = LH
        else:       h = min(max(LH, lines * LH), 180)
        rd.row_dimensions[r].height = h; state["r"] += 1

    def spacer(h):
        rd.row_dimensions[state["r"]].height = h; state["r"] += 1

    def divider(text):
        spacer(16)
        r = state["r"]; rd.merge_cells(f"A{r}:B{r}")
        c = rd.cell(r, 1); c.value = text; c.font = F_DIV
        c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = CTR_L
        rd.row_dimensions[r].height = 48; state["r"] += 1
        spacer(8)

    # title + instructions
    r = state["r"]; rd.merge_cells(f"A{r}:B{r}")
    c = rd.cell(r, 1); c.value = "TIMELINE READER"; c.font = F_TITLE
    c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = CTR_L
    rd.row_dimensions[r].height = 60; state["r"] += 1
    r = state["r"]; rd.merge_cells(f"A{r}:B{r}")
    c = rd.cell(r, 1)
    c.value = "One card per entry. Press Ctrl+F and type an ID (e.g. S34) to jump straight to it."
    c.font = F_HINT; rd.row_dimensions[r].height = 34; state["r"] += 1
    spacer(12)

    cards = 0
    for ws in chron:
        divider(DIVIDER_LABELS.get(ws.title, ws.title.upper()))
        for row in _rows_of(ws):
            ID, Section, Chapter, Date, Cal, Event, Chars, Cross, Ev, Pas = row
            band(f"{ID}    {Chapter}".rstrip())
            when = Date + (f"   ·   calendar {Cal}" if Cal else "") + (f"   ·   Pass {Pas}" if Pas else "")
            field("When", when, small=True)
            field("Section", Section, small=True)
            field("Event", Event, big=True)
            field("Who", Chars)
            field("Cross-refs", Cross)
            field("Evidence", Ev)
            spacer(12)
            cards += 1

    rd.sheet_view.showGridLines = False
    rd.freeze_panes = "A4"
    wb.save(xlsx_path)
    return cards


if __name__ == "__main__":
    n = build_reader()
    print(f"Reader rebuilt: {n} cards.")
