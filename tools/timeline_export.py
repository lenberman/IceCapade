#!/usr/bin/env python3
"""
timeline_export.py — mirror the timeline workbook into diff-able CSVs.

The canonical timeline lives in analysis/IceCapade_timeline.xlsx (a binary file
git can't diff). This tool emits one CSV per sheet into analysis/timeline/ so that
(a) git produces a readable diff when the timeline changes, and
(b) tools/anchors.py can validate the timeline against the manuscript.

For the two chronology sheets it also resolves each row's free-text Chapter label
to the html/ basename(s) it depends on and writes that as a trailing `ChapterFile`
column. Unresolvable labels are written as `UNRESOLVED:<label>` so the lint in
anchors.py can surface them (e.g. a date-range label like "Spring 100" that names
no chapter). The resolution lives here, in the exporter, so the map has one home;
anchors.py just reads the resolved column.

Usage:
  python3 tools/timeline_export.py            # regenerate analysis/timeline/*.csv
  python3 tools/timeline_export.py --check     # exit 1 if the CSVs are out of date
                                               # (mirror differs from the workbook)
"""
import os, re, csv, sys, glob, argparse, io

ROOT     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX     = os.path.join(ROOT, "analysis", "IceCapade_timeline.xlsx")
OUTDIR   = os.path.join(ROOT, "analysis", "timeline")
HTMLD    = os.path.join(ROOT, "html")

# The chronology sheets carry these columns (col 3 = the free-text chapter label).
CHRON_SHEETS = ("Historical", "Slow Mend")   # matched as a prefix of the sheet title
CHAPTER_COL  = 2                              # 0-based index of "Chapter"

# UI-only sheets (formula-driven views) carry no source data — don't mirror them.
SKIP_SHEETS  = ("Reader",)                    # matched as a prefix of the sheet title

# free-text chapter label  ->  html/ basename. Keys are normalized (lowercased,
# punctuation stripped) so minor title drift ("Distant Sparks" vs
# "Distant Sparks, Converging Paths") still resolves.
TITLE_MAP = {
    "front matter": "FrontMatter", "family tree": "FrontMatter", "key dates": "FrontMatter",
    "relics echoes": "Relics", "relics": "Relics",
    "stinsard university": "Stinsard", "stinsard": "Stinsard",
    "entanglement and consequence": "Entanglement", "entanglement": "Entanglement",
    "reckoning inheritance": "Reckoning", "reckoning": "Reckoning",
    "the long flight east": "LongFlight", "long flight east": "LongFlight",
    "touchdown in a new world": "Touchdown", "touchdown": "Touchdown",
    "distant sparks converging paths": "Distant", "distant sparks": "Distant", "distant": "Distant",
    "bonds of support shifting sands": "Bonds", "bonds of support": "Bonds", "bonds": "Bonds",
    "the middle years": "Middle", "middle years": "Middle",
    "thesis defense": "Thesis", "thesis": "Thesis",
    "the marriage": "Marriage", "marriage": "Marriage",
    "the ghost in the machine": "Ghost", "ghost in the machine": "Ghost", "ghost": "Ghost",
    "early crisis response": "Early", "early": "Early",
    "the day aum awakened the world": "Awakening", "the day aum awakened": "Awakening", "awakening": "Awakening",
    "aum speaks": "Speaks", "speaks": "Speaks",
    "beginning of the end": "Beginning", "beginning": "Beginning",
    "here and now": "Here", "here": "Here",
    "november": "November",
    "preparing the endpoint": "Preparing", "preparing": "Preparing",
    "twins": "TWINS",
    "progress": "Progress",
    "zums education": "Education", "zum s education": "Education", "education": "Education",
    "rescuing kaelen": "Rescuing", "rescuing": "Rescuing",
    "logic vs love": "Logic", "logic vs. love": "Logic", "logic": "Logic",
    "waiting": "Waiting",
    "welcome": "Welcome",
}

def _norm_title(s):
    s = (s or "").lower().replace("&", " ").replace(".", " ")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def resolve_chapter(label, chapset):
    """A label may be composite ('Aum Speaks / Beginning of the End',
    'Front Matter; Reckoning & Inheritance'). Split, map each part, dedupe."""
    if not label:
        return ""
    parts = re.split(r"[;/]", label)
    out, seen = [], set()
    unresolved = []
    for p in parts:
        key = _norm_title(p)
        if not key:
            continue
        base = TITLE_MAP.get(key)
        if base and base in chapset:
            if base not in seen:
                seen.add(base); out.append(base)
        else:
            unresolved.append(p.strip())
    if not out and unresolved:
        return "UNRESOLVED:" + "|".join(unresolved)
    if unresolved:  # partial: keep resolved, flag the rest
        return ";".join(out) + ";UNRESOLVED:" + "|".join(unresolved)
    return ";".join(out)

def sheet_csv_name(title):
    return re.sub(r"[^A-Za-z0-9]+", "_", title).strip("_") + ".csv"

def build_csv_text(ws, is_chron, chapset):
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    header = [c.value if c.value is not None else "" for c in ws[1]]
    if is_chron:
        header = header + ["ChapterFile"]
    w.writerow(header)
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in row):
            continue
        row = ["" if v is None else str(v) for v in row]
        if is_chron:
            label = row[CHAPTER_COL] if len(row) > CHAPTER_COL else ""
            row = row + [resolve_chapter(label, chapset)]
        w.writerow(row)
    return buf.getvalue()

def chapters():
    return {os.path.splitext(os.path.basename(p))[0] for p in glob.glob(os.path.join(HTMLD, "*.html"))}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true",
                    help="exit 1 if analysis/timeline/*.csv differ from the workbook (don't write)")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("timeline_export: openpyxl not installed (pip install openpyxl --break-system-packages)", file=sys.stderr)
        sys.exit(2)
    if not os.path.exists(XLSX):
        print(f"timeline_export: workbook not found at {XLSX}", file=sys.stderr)
        sys.exit(2)

    import warnings; warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    chapset = chapters()
    os.makedirs(OUTDIR, exist_ok=True)

    drift = []
    written = []
    for ws in wb.worksheets:
        if ws.title.startswith(SKIP_SHEETS):
            continue                          # UI-only view; nothing to mirror
        is_chron = ws.title.startswith(CHRON_SHEETS)
        text = build_csv_text(ws, is_chron, chapset)
        path = os.path.join(OUTDIR, sheet_csv_name(ws.title))
        if args.check:
            old = open(path, encoding="utf-8").read() if os.path.exists(path) else None
            if old != text:
                drift.append(os.path.basename(path))
        else:
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
            written.append(os.path.basename(path))

    if args.check:
        if drift:
            print("STALE MIRROR — regenerate with `python3 tools/timeline_export.py`:")
            for d in drift:
                print("  ", d)
            sys.exit(1)
        print("timeline CSV mirror is up to date.")
        sys.exit(0)

    print(f"wrote {len(written)} CSV(s) to analysis/timeline/:")
    for n in written:
        print("  ", n)

if __name__ == "__main__":
    main()
